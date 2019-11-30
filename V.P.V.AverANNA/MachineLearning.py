import pandas as pd
import numpy as np
import openpyxl as xl
import wordfreq as wq
from nltk.stem import PorterStemmer
from sklearn.cluster import MeanShift

# defs
def tfidf (sentence , words,data_f):
    pst = PorterStemmer()
    s_sente = sentence.split()
    for i in range (len(s_sente)):
        s_sente[i]=pst.stem(s_sente[i])
    arr=[]
    for i in range (len(words)):
        try:
            word = float(words[i])
        except :
            if len(words[i])>1:
                arr.append((s_sente.count(words[i])/len(s_sente))*100)
    #print(len(arr),len(words),len(data_f.columns))
    data_f.loc[sentence]= arr
    return(data_f)

# opening excel file
workbook = xl.load_workbook(filename = 'eng-sentences.xlsx')
sheet = workbook['Sheet1']

# create dataframe
dataframe = pd.DataFrame()

# getting words from wordfreq
words = words = wq.top_n_list('en',21000)

# changeing dataFrame
for i in range (len(words)):
    try:
        word = float(words[i])
    except :
        if len(words[i])>1:
            dataframe[words[i]]=[0]

# importing sentences to dataframe
for i in range (1,(sheet.max_row)+1):
    dataframe = tfidf (sheet.cell(row=i,column=1).value , words , dataframe)

#create and fit model
ml_model = MeanShift(bandwidth=1)
ml_model.fit(dataframe)
model_labels = ml_model.labels_
model_cluster_centers = ml_model.cluster_centers_
print(model_labels)
print(model_cluster_centers)
